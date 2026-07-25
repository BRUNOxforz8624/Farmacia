import openpyxl
from datetime import datetime
from typing import List, Dict, Optional, Generator
import re

COLUMN_CONFIG = {
    'barcode': {
        'exact': ['cod. barra', 'cod barra', 'codigo de barras', 'cod barras', 'código de barras'],
        'contains': ['barras', 'barcode', 'ean', 'cod barra']
    },
    'name': {
        'exact': ['descripción', 'descripcion', 'description', 'detalle'],
        'contains': ['descripcion', 'descripción', 'description', 'detalle', 'producto']
    },
    'expiration': {
        'exact': ['fecha venc.', 'fecha vencimiento', 'fec lote', 'fecha lote', 'vencimiento', 'fec. venc.', 'fec venc'],
        'contains': ['vencimiento']
    },
    'quantity': {
        'exact': ['existencia', 'inventario', 'cantidad solicitada', 'pedido', 'stock', 'disponible'],
        'contains': ['existencia', 'inventario', 'stock']
    },
    'price': {
        'exact': ['precio', 'precio unit', 'precio unitario', 'precio uni'],
        'contains': ['precio']
    },
    'supplier': {
        'exact': ['proveedor', 'laboratorio', 'lab', 'fabricante'],
        'contains': ['proveedor', 'laboratorio', 'fabricante']
    },
    'conditions': {
        'exact': ['condición', 'condicion', 'acuerdo comercial', 'dcto. nena', 'dcto. ct', 'dcto. en factura', 'oferta', 'descuento', 'dcto nena'],
        'contains': ['condicion', 'condición', 'acuerdo', 'dcto', 'descuento', 'oferta']
    }
}

def normalize(text: str) -> str:
    text = text.lower().strip()
    text = re.sub(r'[\s]+', ' ', text)
    text = text.replace('(', '').replace(')', '').replace('$', '')
    text = text.replace('á', 'a').replace('é', 'e').replace('í', 'i').replace('ó', 'o').replace('ú', 'u')
    text = text.replace('.', '')
    return text.strip()

def parse_excel(file_path: str) -> Generator[Dict, None, None]:
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            yield from process_sheet(sheet)
        wb.close()
    except Exception as e:
        raise Exception(f"Error leyendo Excel: {str(e)}")

def process_sheet(sheet) -> Generator[Dict, None, None]:
    col_map = {}
    data_start = 0
    
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
        if row_idx >= 50:
            break
        if not row or all(cell is None for cell in row):
            continue
        
        non_empty = [h for h in row if h]
        if len(non_empty) < 3:
            continue
        
        headers = [str(h).strip() if h else '' for h in row]
        candidate_map = map_columns(headers)
        
        if candidate_map and len(candidate_map) >= 3:
            unique_cols = set(candidate_map.values())
            if len(unique_cols) >= 3:
                col_map = candidate_map
                data_start = row_idx + 1
                break
    
    if not col_map:
        return
    
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
        if row_idx < data_start:
            continue
        if not row or all(cell is None for cell in row):
            continue
        
        product = extract_product_data(row, col_map)
        if product:
            yield product

def map_columns(headers: List[str]) -> Dict[str, int]:
    col_map = {}
    headers_norm = [normalize(h) for h in headers]
    
    print(f"[EXCEL DEBUG] Headers originales: {headers}")
    print(f"[EXCEL DEBUG] Headers normalizados: {headers_norm}")
    
    for field, config in COLUMN_CONFIG.items():
        for keyword in config['exact']:
            kw_norm = normalize(keyword)
            for i, h in enumerate(headers_norm):
                if h == kw_norm:
                    col_map[field] = i
                    print(f"[EXCEL DEBUG] Columna '{field}' = col {i} ('{headers[i]}')")
                    break
            if field in col_map:
                break
    
    for field, config in COLUMN_CONFIG.items():
        if field in col_map:
            continue
        for keyword in config['contains']:
            kw_norm = normalize(keyword)
            for i, h in enumerate(headers_norm):
                if not h or len(h) < 3:
                    continue
                if kw_norm in h:
                    col_map[field] = i
                    print(f"[EXCEL DEBUG] Columna '{field}' = col {i} ('{headers[i]}' via contains)")
                    break
            if field in col_map:
                break
    
    print(f"[EXCEL DEBUG] Columnas mapeadas: {col_map}")
    return col_map

def extract_product_data(row: tuple, col_map: Dict) -> Optional[Dict]:
    try:
        barcode = get_cell(row, col_map, 'barcode')
        name = get_cell(row, col_map, 'name')
        expiration = parse_date(get_cell(row, col_map, 'expiration'))
        quantity = parse_integer(get_cell(row, col_map, 'quantity'))
        price = parse_price(get_cell(row, col_map, 'price'))
        supplier = get_cell(row, col_map, 'supplier')
        conditions = get_cell(row, col_map, 'conditions')
        
        if not name or price is None:
            return None
        
        return {
            'barcode': barcode,
            'name': name,
            'expiration_date': expiration,
            'quantity': quantity or 0,
            'price': price,
            'supplier': supplier,
            'special_conditions': conditions
        }
    except Exception:
        return None

def get_cell(row: tuple, col_map: Dict, field: str) -> Optional[str]:
    if field in col_map and col_map[field] < len(row):
        value = row[col_map[field]]
        return clean_text(value)
    return None

def clean_text(text) -> str:
    if text is None:
        return ''
    if isinstance(text, (int, float)):
        return str(text)
    text = str(text).strip()
    text = re.sub(r'\s+', ' ', text)
    return text

def parse_price(value) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    try:
        text = str(value).strip()
        text = text.replace('$', '').replace(' ', '')
        if ',' in text and '.' in text:
            text = text.replace('.', '').replace(',', '.')
        elif ',' in text:
            text = text.replace(',', '.')
        text = re.sub(r'[^\d.]', '', text)
        if text:
            return float(text)
    except:
        pass
    return None

def parse_integer(value) -> Optional[int]:
    if value is None:
        return None
    if isinstance(value, int):
        return value
    try:
        text = str(value).strip()
        text = re.sub(r'[^\d-]', '', text)
        if text:
            return int(float(text))
    except:
        pass
    return None

def parse_date(value) -> Optional[datetime]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if hasattr(value, 'date'):
        return value.date()
    
    date_formats = [
        '%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d', '%Y/%m/%d',
        '%d/%m/%y', '%d-%m-%y', '%m/%d/%Y', '%m-%d-%Y',
        '%d.%m.%Y', '%d.%m.%y', '%Y%m%d', '%m/%Y', '%m-%Y'
    ]
    text = str(value).strip()
    for fmt in date_formats:
        try:
            return datetime.strptime(text, fmt).date()
        except:
            continue
    return None
