import sys
import os
sys.path.insert(0, os.path.dirname(__file__))

from services.excel_parser import parse_excel, map_columns
import openpyxl
import glob

files = glob.glob(os.path.join(os.path.dirname(__file__), '..', 'uploads', '*.xlsx'))

if not files:
    print("No hay archivos Excel en uploads/")
else:
    for f in files[:3]:
        print(f"\n{'='*60}")
        print(f"Archivo: {os.path.basename(f)}")
        print(f"{'='*60}")
        
        wb = openpyxl.load_workbook(f, read_only=True)
        for name in wb.sheetnames:
            sheet = wb[name]
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i < 12:
                    headers = [str(h).strip() if h else '' for h in row]
                    col_map = map_columns(headers)
                    if col_map:
                        print(f"\n  Headers en fila {i}:")
                        print(f"  Columnas mapeadas: {col_map}")
                    else:
                        print(f"  Fila {i}: (sin match)")
        wb.close()
        
        print(f"\n  Probando parser completo...")
        try:
            products = parse_excel(f)
            print(f"  Productos encontrados: {len(products)}")
            if products:
                print(f"  Primer producto: {products[0]}")
                print(f"  Ultimo producto: {products[-1]}")
        except Exception as e:
            print(f"  ERROR: {e}")
