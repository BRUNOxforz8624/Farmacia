from flask import Blueprint, request, jsonify, send_file
from services.price_comparator import get_best_prices, compare_product
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import io
from datetime import datetime

compare_bp = Blueprint('compare', __name__)

@compare_bp.route('/best', methods=['GET'])
def best_prices():
    min_qty = request.args.get('min_qty', 5, type=int)
    min_months = request.args.get('min_months', 6, type=int)
    
    results = get_best_prices(min_qty, min_months)
    
    return jsonify({
        'filters': {
            'min_quantity': min_qty,
            'min_months': min_months
        },
        'results': results,
        'total': len(results)
    })

@compare_bp.route('/product', methods=['GET'])
def compare_prices():
    product_name = request.args.get('q', '')
    
    if not product_name:
        return jsonify({'error': 'Se requiere parametro q (nombre del producto)'}), 400
    
    min_qty = request.args.get('min_qty', 5, type=int)
    min_months = request.args.get('min_months', 6, type=int)
    
    results = compare_product(product_name, min_qty, min_months)
    
    return jsonify({
        'query': product_name,
        'filters': {
            'min_quantity': min_qty,
            'min_months': min_months
        },
        'results': results,
        'total': len(results)
    })

@compare_bp.route('/cheapest', methods=['GET'])
def cheapest():
    min_qty = request.args.get('min_qty', 5, type=int)
    min_months = request.args.get('min_months', 6, type=int)
    
    results = get_best_prices(min_qty, min_months)
    
    if not results:
        return jsonify({'message': 'No se encontraron productos que cumplan los filtros'})
    
    cheapest = results[0]
    
    return jsonify(cheapest)

@compare_bp.route('/export', methods=['GET'])
def export_excel():
    min_qty = request.args.get('min_qty', 5, type=int)
    min_months = request.args.get('min_months', 6, type=int)
    search = request.args.get('q', '')
    
    if search:
        results = compare_product(search, min_qty, min_months)
    else:
        results = get_best_prices(min_qty, min_months)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resultados"
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    
    headers = ["Codigo", "Producto", "Proveedor", "Precio", "Cantidad", "Vencimiento", "Condicion", "Estado"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    for row, item in enumerate(results, 2):
        ws.cell(row=row, column=1, value=item.get('barcode', '-'))
        ws.cell(row=row, column=2, value=item.get('product_name', ''))
        ws.cell(row=row, column=3, value=item.get('supplier_name', ''))
        ws.cell(row=row, column=4, value=item.get('price', 0))
        ws.cell(row=row, column=5, value=item.get('quantity', 0))
        ws.cell(row=row, column=6, value=item.get('months_until_expiration', 'N/A'))
        ws.cell(row=row, column=7, value=item.get('special_conditions', '-'))
        ws.cell(row=row, column=8, value='Valida' if item.get('is_valid') else 'Rechazada')
    
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 12
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"farmapp_resultados_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@compare_bp.route('/purchase-orders', methods=['GET'])
def purchase_orders():
    min_qty = request.args.get('min_qty', 5, type=int)
    min_months = request.args.get('min_months', 6, type=int)
    search = request.args.get('q', '')
    
    if search:
        results = compare_product(search, min_qty, min_months)
    else:
        results = get_best_prices(min_qty, min_months)
    
    valid = [r for r in results if r.get('is_valid')]
    
    if not valid:
        return jsonify({'error': 'No hay productos validos para generar ordenes'}), 400
    
    by_supplier = {}
    for item in valid:
        name = item['supplier_name']
        if name not in by_supplier:
            by_supplier[name] = []
        by_supplier[name].append(item)
    
    wb = openpyxl.Workbook()
    
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    supplier_font = Font(bold=True, size=13, color="1E293B")
    total_font = Font(bold=True, size=11)
    total_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
    money_format = '#,##0.00'
    
    summary_ws = wb.active
    summary_ws.title = "Resumen"
    
    summary_headers = ["Proveedor", "Productos", "Subtotal"]
    for col, h in enumerate(summary_headers, 1):
        cell = summary_ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    summary_row = 2
    grand_total = 0
    
    for supplier_name in sorted(by_supplier.keys()):
        items = by_supplier[supplier_name]
        ws = wb.create_sheet(title=supplier_name[:31])
        
        ws.cell(row=1, column=1, value=f"Orden de Compra - {supplier_name}")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        
        ws.cell(row=2, column=1, value=f"Fecha: {datetime.now().strftime('%d/%m/%Y')}")
        ws.cell(row=2, column=1).font = Font(color="64748B")
        
        headers = ["#", "Codigo Barra", "Producto", "Droguería", "Precio Unit.", "Cantidad", "Subtotal", "Condicion"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
        
        supplier_total = 0
        for i, item in enumerate(items, 1):
            qty = item['quantity']
            price = item['price']
            subtotal = qty * price
            supplier_total += subtotal
            
            row = 4 + i
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=item.get('barcode', '-'))
            ws.cell(row=row, column=3, value=item['product_name'])
            ws.cell(row=row, column=4, value=supplier_name)
            ws.cell(row=row, column=5, value=price).number_format = money_format
            ws.cell(row=row, column=6, value=qty)
            ws.cell(row=row, column=7, value=subtotal).number_format = money_format
            ws.cell(row=row, column=8, value=item.get('special_conditions', '-'))
        
        total_row = 4 + len(items) + 1
        ws.cell(row=total_row, column=3, value="TOTAL").font = total_font
        ws.cell(row=total_row, column=3).fill = total_fill
        ws.cell(row=total_row, column=7, value=supplier_total).font = total_font
        ws.cell(row=total_row, column=7).fill = total_fill
        ws.cell(row=total_row, column=7).number_format = money_format
        
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 14
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 14
        ws.column_dimensions['H'].width = 25
        
        summary_ws.cell(row=summary_row, column=1, value=supplier_name)
        summary_ws.cell(row=summary_row, column=2, value=len(items))
        summary_ws.cell(row=summary_row, column=3, value=supplier_total).number_format = money_format
        
        grand_total += supplier_total
        summary_row += 1
    
    summary_row += 1
    summary_ws.cell(row=summary_row, column=1, value="TOTAL GENERAL").font = total_font
    summary_ws.cell(row=summary_row, column=1).fill = total_fill
    summary_ws.cell(row=summary_row, column=2, value=len(valid)).font = total_font
    summary_ws.cell(row=summary_row, column=2).fill = total_fill
    summary_ws.cell(row=summary_row, column=3, value=grand_total).font = total_font
    summary_ws.cell(row=summary_row, column=3).fill = total_fill
    summary_ws.cell(row=summary_row, column=3).number_format = money_format
    
    summary_ws.column_dimensions['A'].width = 30
    summary_ws.column_dimensions['B'].width = 12
    summary_ws.column_dimensions['C'].width = 15
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"ordenes_compra_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )
